"""CSV/Excel export for screening results."""

import csv
import os
from datetime import date


def _flat_row(r):
    """Convert a nested result dict to a flat row for export."""
    eps = r.get("eps_analysis", {})
    roe = r.get("roe_analysis", {})
    fcf = r.get("fcf_analysis", {})
    bal = r.get("balance_analysis", {})
    div = r.get("dividend_analysis", {})
    dcf = r.get("dcf_analysis", {})

    rev = r.get("revenue_analysis", {})

    return {
        "Symbol": r.get("symbol"),
        "Name": r.get("name"),
        "Sector": r.get("sector"),
        "Industry": r.get("industry"),
        "Market Cap ($B)": r.get("market_cap_b"),
        "Price": r.get("current_price"),
        "P/E": r.get("trailing_pe"),
        "Fundamental Score": r.get("fundamental_score"),
        "EPS Score": eps.get("eps_score"),
        "EPS CAGR (%)": eps.get("eps_growth_rate"),
        "EPS Consistent": "Yes" if eps.get("eps_consistent") else "No",
        "ROE Score": roe.get("roe_score"),
        "ROE (%)": roe.get("roe"),
        "Debt/Equity": roe.get("debt_to_equity"),
        "FCF Score": fcf.get("fcf_score"),
        "FCF ($B)": fcf.get("fcf_current"),
        "FCF Yield (%)": fcf.get("fcf_yield"),
        "FCF Growing": "Yes" if fcf.get("fcf_growing") else "No",
        "Balance Score": bal.get("balance_score"),
        "Current Ratio": bal.get("current_ratio"),
        "Cash/Debt": bal.get("cash_to_debt"),
        "Goodwill (%)": bal.get("goodwill_pct"),
        "Dividend Score": div.get("dividend_score"),
        "Dividend Yield (%)": div.get("dividend_yield_pct"),
        "Payout Ratio (%)": div.get("payout_ratio_pct"),
        "Consecutive Increases": div.get("consecutive_increases"),
        "Revenue CAGR (%)": rev.get("revenue_cagr"),
        "Revenue Growing": "Yes" if rev.get("revenue_growing") else "No",
        "Intrinsic Value": dcf.get("intrinsic_value"),
        "Margin of Safety (%)": dcf.get("margin_of_safety"),
        "Undervalued": "Yes" if dcf.get("undervalued") else "No",
    }


def export_csv(results, filepath=None):
    """Export results to CSV file.

    Args:
        results: List of nested result dicts from screen_stock().
        filepath: Output path. Defaults to scores_YYYY-MM-DD.csv.

    Returns:
        Path to the written file.
    """
    if not results:
        print("  No results to export.")
        return None

    if filepath is None:
        filepath = f"scores_{date.today().isoformat()}.csv"

    rows = [_flat_row(r) for r in results]
    fieldnames = list(rows[0].keys())

    with open(filepath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"  Exported {len(rows)} stocks to {filepath}")
    return filepath


def export_excel(results, filepath=None):
    """Export results to Excel (.xlsx) file.

    Requires openpyxl. Falls back to CSV if not installed.

    Returns:
        Path to the written file.
    """
    if not results:
        print("  No results to export.")
        return None

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
    except ImportError:
        print("  openpyxl not installed. Falling back to CSV export.")
        print("  Install with: pip install openpyxl")
        return export_csv(results, filepath.replace(".xlsx", ".csv") if filepath else None)

    if filepath is None:
        filepath = f"scores_{date.today().isoformat()}.xlsx"

    rows = [_flat_row(r) for r in results]
    fieldnames = list(rows[0].keys())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Fundamental Scores {date.today().isoformat()}"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    for col, name in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(fieldnames, 1):
            val = row[key]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(filepath)
    print(f"  Exported {len(rows)} stocks to {filepath}")
    return filepath


def export_csv_from_db(tickers=None):
    """Export latest scores from the database to a dated CSV snapshot.

    Args:
        tickers: Optional list of ticker symbols to include.
                 If None, exports all tickers in the database.

    Returns:
        Path to the written file.
    """
    from utils.scores_db import get_latest_scores

    latest = get_latest_scores()
    if not latest:
        print("  No scores in database to export.")
        return None

    if tickers:
        t_set = set(t.upper() for t in tickers)
        latest = [r for r in latest if r["symbol"] in t_set]

    if not latest:
        print("  No matching scores found in database.")
        return None

    filepath = f"snapshot_{date.today().isoformat()}.csv"

    # Use DB column names directly — skip internal columns
    skip = {"id"}
    fieldnames = [k for k in latest[0].keys() if k not in skip]

    with open(filepath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(latest)

    print(f"  📁 Exported {len(latest)} stocks to {filepath}")
    return filepath
